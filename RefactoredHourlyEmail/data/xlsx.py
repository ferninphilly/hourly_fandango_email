__author__ = 'fpombeiro'
import openpyxl
import yaml
from datetime import datetime
from config import TODAY, YAML_LOCATION
import os

f = open(YAML_LOCATION + 'hourly.yaml')
dataMap = yaml.safe_load(f)

def report_data_excel(sourcefile):
    return openpyxl.load_workbook(filename=sourcefile).worksheets[0]

def baseline_month_excel(worksheet,dataobj=dict()):
    dataobj['baseline'] = dict()
    dataobj['baseline_month'] = dict()
    for row in worksheet.iter_rows():
        if row[0].value is not None:
            if datetime.date(row[0].value) == TODAY:
                dataobj['baseline']= {
                        'Date' : datetime.date(row[0].value),
                        'Day_of_week' : row[1].value,
                        'Total_count' : int(row[2].value)
                }
                dataobj['baseline_month'][row[0].value] = {
                        'Date': datetime.date(row[0].value),
                        'Day_of_week': row[1].value,
                        'Total_count': int(row[2].value)
                }
            elif datetime.date(row[0].value).month == TODAY.month \
                    and datetime.date(row[0].value).year == TODAY.year:
                dataobj['baseline_month'][row[0].value] = {
                    'Date': datetime.date(row[0].value),
                    'Day_of_week': row[1].value,
                    'Total_count': int(row[2].value)
                }
    return dataobj

def smoothing_excel(worksheet, dataobj=dict()):
    for row in worksheet.iter_rows(row_offset=1):
        if row[1].value == TODAY.strftime('%A'):
            if row[1].value not in dataobj:
                dataobj[row[1].value] = dict()
                dataobj[row[1].value] = {row[0].value : row[2].value}
            else:
                dataobj[row[1].value].update({row[0].value : row[2].value})
    return dataobj

def get_excel_reports():
    dataMapsub = dataMap['hourly_report']['file_locations']
    monthly_excel = report_data_excel(dataMapsub['baseline'])
    monthly_baseline = baseline_month_excel(monthly_excel)
    hourly_baseline = smoothing_excel(report_data_excel(dataMapsub['hourlysmooth']))
    return dict(monthly_baseline, **hourly_baseline)

